from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Producto
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

@login_required
def generar_reporte_excel(request):
    # Crear un nuevo libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Estilo para los encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    # Encabezados
    headers = ['ID', 'Nombre', 'Descripción', 'Precio', 'Stock', 'Categoría']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    # Obtener todos los productos
    productos = Producto.objects.all()

    # Llenar datos
    for row, producto in enumerate(productos, 2):
        ws.cell(row=row, column=1, value=producto.id)
        ws.cell(row=row, column=2, value=producto.nombre)
        ws.cell(row=row, column=3, value=producto.descripcion)
        ws.cell(row=row, column=4, value=float(producto.precio))
        ws.cell(row=row, column=5, value=producto.stock)
        ws.cell(row=row, column=6, value=str(producto.categoria))

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Generar nombre de archivo con timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"inventario_{timestamp}.xlsx"

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Guardar el libro de Excel en la respuesta
    wb.save(response)
    
    messages.success(request, "Reporte de inventario generado exitosamente")
    return response

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if not username or not password:
            messages.error(request, "Por favor complete todos los campos")
            return redirect('login')
            
        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, f"Bienvenido {user.username}!")
            return redirect('home')
        else:
            messages.error(request, "Credenciales incorrectas. Por favor intente nuevamente.")
            return redirect('login')
    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    messages.success(request, "Has cerrado sesión exitosamente")
    return redirect('login')

def registro_view(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, "¡Registro exitoso! Por favor inicia sesión.")
            return redirect('login')
        else:
            for msg in form.error_messages:
                messages.error(request, form.error_messages[msg])
    return render(request, 'registro.html', {'form': UserCreationForm()})

@login_required
def checkout_view(request):
    if request.method == 'POST':
        # Lógica de procesamiento de compra aquí
        # ...
        messages.success(request, "¡Compra realizada exitosamente! Tu pedido está en proceso.")
        return redirect('pedidos')
    return render(request, 'checkout.html')