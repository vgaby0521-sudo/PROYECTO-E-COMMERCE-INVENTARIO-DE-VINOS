from decimal import Decimal, ROUND_HALF_UP
import json
from functools import wraps
from datetime import datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
from django.contrib import messages

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render
from .models import Inventario
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.decorators import user_passes_test

from django.contrib.auth.decorators import login_required, user_passes_test
from functools import wraps
from django.shortcuts import redirect
from django.contrib import messages

def admin_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not request.user.is_superuser:
            messages.error(request, 'Acceso denegado. Solo administradores.')
            return redirect('home')
        return view_func(request, *args, **kwargs)
    return wrapper


from .models import (
    Categoria, Producto, Perfil, CarritoItem, Pedido, DetallePedido,
    Factura, Notificacion, Recompensa, Inventario, Proveedor
)



def _calc_totals_from_items(items):

    total = Decimal('0.00')
    for item in items:
        try:
            total += Decimal(str(item.subtotal))
        except Exception:
            # si item.subtotal falla, lo omitimos (no debería ocurrir)
            continue
    total = total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    impuestos = (total * Decimal('0.19')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    total_final = (total + impuestos).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    return total, impuestos, total_final


def home(request):
    categorias = Categoria.objects.all()
    productos_destacados = Producto.objects.filter(activo=True).order_by('-fecha_agregado')[:6]
    ofertas = Producto.objects.filter(activo=True, precio_oferta__isnull=False).order_by('-fecha_agregado')[:6]

    context = {
        'categorias': categorias,
        'productos_destacados': productos_destacados,
        'ofertas': ofertas,
    }
    return render(request, 'home.html', context)


def catalogo(request):
    productos = Producto.objects.filter(activo=True)
    categorias = Categoria.objects.all()

    categoria_id = request.GET.get('categoria')
    busqueda = request.GET.get('busqueda')

    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)

    if busqueda:
        productos = productos.filter(
            Q(nombre__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )

    context = {
        'productos': productos,
        'categorias': categorias,
        'categoria_seleccionada': categoria_id,
        'busqueda': busqueda,
    }
    return render(request, 'catalogo.html', context)


def detalle_producto(request, *args, **kwargs):
    """Mostrar detalle de un producto.

    Esta función acepta tanto el keyword 'producto_id' (desde la URL)
    como 'id' para compatibilidad con diferentes firmas. No modifica
    otras rutas ni plantillas: extrae el id disponible y carga el producto.
    """
    # Aceptar ambos nombres de parámetro para evitar TypeError si la URL
    # o la vista usan distinto nombre ('producto_id' o 'id').
    producto_id = kwargs.get('producto_id') or kwargs.get('id')
    # Si no viene por kwargs, intentar obtenerlo por posición (p. ej. /producto/1/)
    if producto_id is None and args:
        try:
            producto_id = int(args[0])
        except Exception:
            producto_id = None

    if producto_id is None:
        return HttpResponse('Producto no especificado', status=400)

    producto = get_object_or_404(Producto, id=producto_id)
    relacionados = Producto.objects.filter(
        categoria=producto.categoria,
        activo=True
    ).exclude(id=producto_id)[:4]

    context = {
        'producto': producto,
        'relacionados': relacionados,
    }
    return render(request, 'detalle_producto.html', context)


@login_required(login_url='login')
def carrito(request):
    if request.user.is_superuser:
        messages.error(request, "Los administradores no tienen acceso al carrito de compras")
        return redirect('admin_panel')
        
    items = CarritoItem.objects.filter(usuario=request.user)
    total, impuestos, total_final = _calc_totals_from_items(items)

    context = {
        'items': items,
        'total': total,
        'impuestos': impuestos,
        'total_final': total_final,
    }
    return render(request, 'carrito.html', context)


@login_required(login_url='login')
@require_http_methods(["POST"])
def agregar_carrito_api(request):
    if request.user.is_superuser:
        return JsonResponse({
            'success': False,
            'message': 'Los administradores no tienen acceso al carrito de compras'
        }, status=403)
        
    try:
        data = json.loads(request.body)
        producto_id = data.get('producto_id')
        cantidad = int(data.get('cantidad', 1))

        producto = get_object_or_404(Producto, id=producto_id)

        item, created = CarritoItem.objects.get_or_create(
            usuario=request.user,
            producto=producto,
            defaults={'cantidad': cantidad}
        )

        if not created:
            item.cantidad += cantidad
            item.save()

        return JsonResponse({
            'success': True,
            'message': f'{producto.nombre} agregado al carrito'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

@login_required(login_url='login')
@require_http_methods(["POST"])
def agregar_carrito(request, producto_id):
    if request.user.is_superuser:
        return JsonResponse({
            'success': False,
            'message': 'Los administradores no tienen acceso al carrito de compras'
        }, status=403)
        
    try:
        cantidad = 1  # Por defecto se agrega 1 unidad
        if request.method == 'POST':
            data = json.loads(request.body)
            cantidad = int(data.get('cantidad', 1))

        producto = get_object_or_404(Producto, id=producto_id)

        item, created = CarritoItem.objects.get_or_create(
            usuario=request.user,
            producto=producto,
            defaults={'cantidad': cantidad}
        )

        if not created:
            item.cantidad += cantidad
            item.save()

        return JsonResponse({
            'success': True,
            'message': f'{producto.nombre} agregado al carrito'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


@login_required(login_url='login')
@require_http_methods(["POST"])
def eliminar_carrito(request, item_id):
    if request.user.is_superuser:
        return JsonResponse({
            'success': False,
            'message': 'Los administradores no tienen acceso al carrito de compras'
        }, status=403)
        
    try:
        item = get_object_or_404(CarritoItem, id=item_id, usuario=request.user)
        item.delete()
        return JsonResponse({'success': True, 'message': 'Producto removido del carrito'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


@login_required(login_url='login')
@require_http_methods(["POST"])
def actualizar_carrito(request, item_id):
    if request.user.is_superuser:
        return JsonResponse({
            'success': False,
            'message': 'Los administradores no tienen acceso al carrito de compras'
        }, status=403)
        
    try:
        data = json.loads(request.body)
        cantidad = int(data.get('cantidad', 1))

        item = get_object_or_404(CarritoItem, id=item_id, usuario=request.user)

        if cantidad <= 0:
            item.delete()
            message = 'Producto removido'
        else:
            item.cantidad = cantidad
            item.save()
            message = 'Carrito actualizado'

        return JsonResponse({'success': True, 'message': message})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


@login_required(login_url='login')
def checkout(request):
    if request.user.is_superuser:
        messages.error(request, "Los administradores no tienen acceso al proceso de compra")
        return redirect('admin_panel')
        
    items = CarritoItem.objects.filter(usuario=request.user)
    total, impuestos, total_final = _calc_totals_from_items(items)

    context = {
        'items': items,
        'total': total,
        'impuestos': impuestos,
        'total_final': total_final,
    }
    return render(request, 'checkout.html', context)


@login_required(login_url='login')
@require_POST
def crear_pedido(request):
    if request.user.is_superuser:
        return JsonResponse({
            'success': False,
            'message': 'Los administradores no tienen acceso al proceso de compra'
        }, status=403)
        
    try:
        data = json.loads(request.body or "{}")
        items = CarritoItem.objects.filter(usuario=request.user)

        if not items.exists():
            return JsonResponse({'success': False, 'message': 'Carrito vacío'}, status=400)

        total, impuestos, total_final = _calc_totals_from_items(items)

        pedido = Pedido.objects.create(
            cliente=request.user,
            total=total,
            metodo_pago=data.get('metodo_pago', 'tarjeta'),
            direccion_envio=data.get('direccion', ''),
            ciudad_envio=data.get('ciudad', ''),
            telefono_envio=data.get('telefono', ''),
            observaciones=data.get('observaciones', '')
        )

        for item in items:
            DetallePedido.objects.create(
                pedido=pedido,
                producto=item.producto,
                cantidad=item.cantidad,
                precio_unitario=item.producto.precio_final
            )

            # actualizar stock y guardar movimiento de inventario
            cantidad_anterior = item.producto.stock
            item.producto.stock = max(0, item.producto.stock - item.cantidad)
            item.producto.save()

            Inventario.objects.create(
                producto=item.producto,
                cantidad_anterior=cantidad_anterior,
                cantidad_nueva=item.producto.stock,
                tipo_movimiento='salida',
                motivo=f'Venta Pedido #{pedido.id}',
                responsable=request.user
            )

        Factura.objects.create(
            pedido=pedido,
            numero_factura=f'FAC-{pedido.id:06d}',
            total=total
        )

        puntos_ganados = int((total / Decimal('10')).to_integral_value(rounding=ROUND_HALF_UP))
        Recompensa.objects.create(
            usuario=request.user,
            puntos=puntos_ganados,
            motivo=f'Compra Pedido #{pedido.id}',
            tipo='ganado',
            pedido=pedido
        )

        # Asegurar que el perfil exista
        perfil = getattr(request.user, 'perfil', None)
        if perfil is None:
            perfil = Perfil.objects.create(
                usuario=request.user,
                nombre_completo=request.user.get_full_name() or request.user.username
            )
        perfil.puntos_recompensa = (perfil.puntos_recompensa or 0) + puntos_ganados
        perfil.save()

        Notificacion.objects.create(
            usuario=request.user,
            titulo='Pedido Confirmado',
            mensaje=f'Tu pedido #{pedido.id} ha sido confirmado',
            tipo='exito'
        )

        items.delete()

        return JsonResponse({
            'success': True,
            'pedido_id': pedido.id,
            'message': 'Pedido creado exitosamente'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


@login_required(login_url='login')
def mis_pedidos(request):
    pedidos = Pedido.objects.filter(cliente=request.user)

    context = {
        'pedidos': pedidos,
    }
    return render(request, 'mis_pedidos.html', context)


@login_required(login_url='login')
def detalle_pedido(request, id):
    pedido = get_object_or_404(Pedido, id=id, cliente=request.user)
    detalles = pedido.detalles.all()

    context = {
        'pedido': pedido,
        'detalles': detalles,
    }
    return render(request, 'detalle_pedido.html', context)


@login_required(login_url='login')
def perfil_usuario(request):
    try:
        perfil = request.user.perfil
    except Perfil.DoesNotExist:
        tipo = 'admin' if request.user.is_superuser else 'cliente'
        perfil = Perfil.objects.create(
            usuario=request.user,
            nombre_completo=request.user.get_full_name() or request.user.username,
            tipo_usuario=tipo
        )

    notificaciones = Notificacion.objects.filter(usuario=request.user).order_by('-fecha')[:5]
    recompensas = Recompensa.objects.filter(usuario=request.user)

    context = {
        'perfil': perfil,
        'notificaciones': notificaciones,
        'recompensas': recompensas,
        'is_admin': request.user.is_superuser
    }
    return render(request, 'perfil_usuario.html', context)


@login_required(login_url='login')
@require_POST
def actualizar_perfil(request):
    try:
        data = json.loads(request.body or "{}")
        perfil = getattr(request.user, 'perfil', None)
        if perfil is None:
            perfil = Perfil.objects.create(usuario=request.user)

        perfil.nombre_completo = data.get('nombre_completo', perfil.nombre_completo)
        perfil.telefono = data.get('telefono', perfil.telefono)
        perfil.direccion = data.get('direccion', perfil.direccion)
        perfil.ciudad = data.get('ciudad', perfil.ciudad)
        perfil.save()

        return JsonResponse({'success': True, 'message': 'Perfil actualizado'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


def login_view(request):
    next_url = request.GET.get('next')
    
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return redirect('admin_panel')
        return redirect('home')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        if not username or not password:
            messages.error(request, "⚠️ Por favor complete todos los campos")
            return redirect('login')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, f"✅ ¡Bienvenido {user.username}!")
            if next_url:
                return redirect(next_url)
            if user.is_superuser:
                return redirect('admin_panel')
            return redirect('home')
        else:
            messages.error(request, "❌ Credenciales incorrectas. Por favor intente nuevamente.")
            return redirect('login')

    return render(request, 'login.html')


def registro_view(request):
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        password_confirm = request.POST.get('password_confirm')
        nombre_completo = request.POST.get('nombre_completo')
        telefono = request.POST.get('telefono')
        direccion = request.POST.get('direccion')

        if not all([username, email, password, password_confirm]):
            messages.error(request, "⚠️ Por favor complete todos los campos obligatorios")
            return render(request, 'registro.html')

        if password != password_confirm:
            messages.error(request, "❌ Las contraseñas no coinciden")
            return render(request, 'registro.html')

        if User.objects.filter(username=username).exists():
            messages.error(request, "❌ El nombre de usuario ya está en uso")
            return render(request, 'registro.html')

        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=nombre_completo.split()[0] if nombre_completo else '',
            last_name=' '.join(nombre_completo.split()[1:]) if nombre_completo else ''
        )

        tipo = 'admin' if user.is_superuser else 'cliente'
        perfil = Perfil.objects.create(
                usuario=user,
                nombre_completo=nombre_completo,
                telefono=telefono,
                direccion=direccion,
                tipo_usuario=tipo
            )

        login(request, user)
        return redirect('home')

    return render(request, 'registro.html')


def logout_view(request):
    nombre_usuario = request.user.username
    logout(request)
    messages.success(request, f"✅ Has cerrado sesión exitosamente. ¡Hasta pronto {nombre_usuario}!")
    return redirect('login')


@csrf_exempt
def api_productos(request):
    categoria_id = request.GET.get('categoria')
    busqueda = request.GET.get('busqueda')

    productos = Producto.objects.filter(activo=True)

    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)

    if busqueda:
        productos = productos.filter(
            Q(nombre__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )

    data = []
    for producto in productos:
        data.append({
            'id': producto.id,
            'nombre': producto.nombre,
            'descripcion': producto.descripcion[:100],
            'precio': float(producto.precio),
            'precio_final': float(producto.precio_final),
            'stock': producto.stock,
            'imagen': producto.imagen_url,
        })

    return JsonResponse({'productos': data})


@login_required(login_url='login')
def api_carrito(request):
    if request.user.is_superuser:
        return JsonResponse({'cantidad_items': 0})
        
    cantidad_items = CarritoItem.objects.filter(usuario=request.user).count()
    return JsonResponse({'cantidad_items': cantidad_items})


@admin_required
@admin_required
def generar_reporte_excel(request):
    try:
        # Crear un nuevo libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario de Vinos"

        # Estilos
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="722F37", end_color="722F37", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )

        # Título del reporte
        ws.merge_cells('A1:H1')
        ws['A1'] = "Reporte de Inventario - Wine Shop"
        ws['A1'].font = Font(name='Arial', size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center")

        # Fecha del reporte
        ws.merge_cells('A2:H2')
        ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A2'].font = Font(name='Arial', size=10)
        ws['A2'].alignment = Alignment(horizontal="center")

        # Encabezados
        headers = [
            'ID', 'Producto', 'Descripción', 'Categoría', 'Stock', 'Precio', 
            'Proveedor', 'Última Actualización'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Datos
        productos = Producto.objects.select_related('categoria', 'proveedor').all()
        row = 5
        for producto in productos:
            ws.cell(row=row, column=1, value=producto.id)
            ws.cell(row=row, column=2, value=producto.nombre)
            ws.cell(row=row, column=3, value=producto.descripcion)
            ws.cell(row=row, column=4, value=str(producto.categoria))
            ws.cell(row=row, column=5, value=producto.stock)
            ws.cell(row=row, column=6, value=float(producto.precio))
            ws.cell(row=row, column=7, value=str(producto.proveedor) if producto.proveedor else 'N/A')
            ws.cell(row=row, column=8, value=producto.fecha_agregado.strftime('%d/%m/%Y %H:%M') if producto.fecha_agregado else None)

            # Aplicar bordes a todas las celdas
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = Alignment(vertical="center")

            # Resaltar stock bajo
            if producto.stock <= 10:
                ws.cell(row=row, column=5).fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")

            row += 1

        # Ajustar ancho de columnas
        for column_cells in ws.columns:
            length = 0
            column = None
            for cell in column_cells:
                try:
                    # Ignorar celdas combinadas
                    if cell.coordinate in ws.merged_cells:
                        continue
                    # Solo procesar celdas normales
                    if not hasattr(cell, 'column_letter'):
                        column = openpyxl.utils.get_column_letter(cell.column)
                    else:
                        column = cell.column_letter
                    # Calcular el ancho necesario
                    if cell.value:
                        length = max(length, len(str(cell.value)))
                except:
                    pass
            
            if column and length > 0:
                ws.column_dimensions[column].width = length + 2

        # Crear la respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Inventario_WineShop_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        # Guardar el archivo
        wb.save(response)
        
        messages.success(request, "📊 Reporte de inventario generado exitosamente")
        return response

    except Exception as e:
        messages.error(request, f"❌ Error al generar el reporte: {str(e)}")
        return redirect('admin_panel')

@admin_required
def generar_reporte_excel(request):
    try:
        # Crear un nuevo libro de Excel con múltiples hojas
        wb = openpyxl.Workbook()
        
        # Configurar estilos comunes
        title_font = Font(name='Arial', size=16, bold=True)
        header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="722F37", end_color="722F37", fill_type="solid")
        date_font = Font(name='Arial', size=10)
        align_center = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. HOJA DE PRODUCTOS
        ws_productos = wb.active
        ws_productos.title = "Productos"
        
        ws_productos['A1'] = "INVENTARIO DE PRODUCTOS - WINE SHOP"
        ws_productos['A1'].font = title_font
        ws_productos['A1'].alignment = align_center
        ws_productos['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws_productos['A2'].font = date_font
        ws_productos['A2'].alignment = align_center

        headers_productos = [
            'ID', 'Producto', 'Descripción', 'Categoría', 'Stock', 'Precio', 
            'Precio Oferta', 'Proveedor', 'Graduación', 'País', 'Viña', 
            'Cosecha', 'Estado', 'Última Actualización'
        ]
        
        # Aplicar encabezados para productos
        for col, header in enumerate(headers_productos, 1):
            cell = ws_productos.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border

        # 2. HOJA DE MOVIMIENTOS
        ws_movimientos = wb.create_sheet("Movimientos")
        
        ws_movimientos['A1'] = "MOVIMIENTOS DE INVENTARIO - WINE SHOP"
        ws_movimientos['A1'].font = title_font
        ws_movimientos['A1'].alignment = align_center
        ws_movimientos['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws_movimientos['A2'].font = date_font
        ws_movimientos['A2'].alignment = align_center

        headers_movimientos = [
            'ID', 'Fecha', 'Producto', 'Tipo Movimiento', 'Cantidad Anterior',
            'Cantidad Nueva', 'Diferencia', 'Motivo', 'Responsable'
        ]
        
        # Aplicar encabezados para movimientos
        for col, header in enumerate(headers_movimientos, 1):
            cell = ws_movimientos.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border

        # 3. HOJA DE PROVEEDORES
        ws_proveedores = wb.create_sheet("Proveedores")
        
        ws_proveedores['A1'] = "PROVEEDORES - WINE SHOP"
        ws_proveedores['A1'].font = title_font
        ws_proveedores['A1'].alignment = align_center
        ws_proveedores['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws_proveedores['A2'].font = date_font
        ws_proveedores['A2'].alignment = align_center

        headers_proveedores = [
            'ID', 'Nombre', 'Correo', 'Teléfono', 'Dirección',
            'Productos Activos', 'Última Actualización'
        ]
        
        # Aplicar encabezados para proveedores
        for col, header in enumerate(headers_proveedores, 1):
            cell = ws_proveedores.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border

        # 1. LLENAR DATOS DE PRODUCTOS
        productos = Producto.objects.select_related('categoria', 'proveedor').all()
        row = 5
        for producto in productos:
            # Datos básicos
            ws_productos.cell(row=row, column=1, value=producto.id)
            ws_productos.cell(row=row, column=2, value=producto.nombre)
            ws_productos.cell(row=row, column=3, value=producto.descripcion)
            ws_productos.cell(row=row, column=4, value=str(producto.categoria))
            ws_productos.cell(row=row, column=5, value=producto.stock)
            ws_productos.cell(row=row, column=6, value=float(producto.precio))
            ws_productos.cell(row=row, column=7, value=float(producto.precio_oferta) if producto.precio_oferta else None)
            ws_productos.cell(row=row, column=8, value=str(producto.proveedor) if producto.proveedor else 'No asignado')
            ws_productos.cell(row=row, column=9, value=producto.graduacion)
            ws_productos.cell(row=row, column=10, value=producto.pais)
            ws_productos.cell(row=row, column=11, value=producto.vina)
            ws_productos.cell(row=row, column=12, value=producto.cosecha)
            
            # Estado con formato condicional basado en stock
            estado_cell = ws_productos.cell(row=row, column=13)
            if producto.stock <= 5:
                estado_cell.value = "CRÍTICO"
                estado_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            elif producto.stock <= 10:
                estado_cell.value = "BAJO"
                estado_cell.fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
            else:
                estado_cell.value = "DISPONIBLE"
                estado_cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
            
            # Fecha de creación (no existe 'updated_at' en el modelo)
            ws_productos.cell(row=row, column=14, value=producto.fecha_agregado.strftime('%d/%m/%Y %H:%M') if producto.fecha_agregado else None)
            
            # Aplicar bordes y alineación
            for col in range(1, len(headers_productos) + 1):
                cell = ws_productos.cell(row=row, column=col)
                cell.border = border
                cell.alignment = align_center
            
            row += 1

        # 2. LLENAR DATOS DE MOVIMIENTOS
        movimientos = Inventario.objects.select_related('producto', 'responsable').order_by('-fecha')
        row = 5
        for movimiento in movimientos:
            ws_movimientos.cell(row=row, column=1, value=movimiento.id)
            ws_movimientos.cell(row=row, column=2, value=movimiento.fecha.strftime('%d/%m/%Y %H:%M'))
            ws_movimientos.cell(row=row, column=3, value=movimiento.producto.nombre)
            ws_movimientos.cell(row=row, column=4, value=movimiento.tipo_movimiento)
            ws_movimientos.cell(row=row, column=5, value=movimiento.cantidad_anterior)
            ws_movimientos.cell(row=row, column=6, value=movimiento.cantidad_nueva)
            ws_movimientos.cell(row=row, column=7, value=movimiento.cantidad_nueva - movimiento.cantidad_anterior)
            ws_movimientos.cell(row=row, column=8, value=movimiento.motivo)
            ws_movimientos.cell(row=row, column=9, value=movimiento.responsable.username)
            
            # Aplicar bordes y alineación
            for col in range(1, len(headers_movimientos) + 1):
                cell = ws_movimientos.cell(row=row, column=col)
                cell.border = border
                cell.alignment = align_center
            
            row += 1

        # 3. LLENAR DATOS DE PROVEEDORES
        proveedores = Proveedor.objects.all()
        row = 5
        for proveedor in proveedores:
            productos_activos = Producto.objects.filter(proveedor=proveedor, activo=True).count()
            
            ws_proveedores.cell(row=row, column=1, value=proveedor.id)
            ws_proveedores.cell(row=row, column=2, value=proveedor.nombre)
            ws_proveedores.cell(row=row, column=3, value=proveedor.correo)
            ws_proveedores.cell(row=row, column=4, value=proveedor.telefono)
            ws_proveedores.cell(row=row, column=5, value=proveedor.direccion)
            ws_proveedores.cell(row=row, column=6, value=productos_activos)
            # Proveedor usa 'fecha_registro' en el modelo
            ws_proveedores.cell(row=row, column=7, value=proveedor.fecha_registro.strftime('%d/%m/%Y %H:%M') if proveedor.fecha_registro else None)
            
            # Aplicar bordes y alineación
            for col in range(1, len(headers_proveedores) + 1):
                cell = ws_proveedores.cell(row=row, column=col)
                cell.border = border
                cell.alignment = align_center
            
            row += 1

        # AJUSTAR ANCHO DE COLUMNAS
        for sheet in [ws_productos, ws_movimientos, ws_proveedores]:
            if sheet == ws_productos:
                headers = headers_productos
            elif sheet == ws_movimientos:
                headers = headers_movimientos
            else:
                headers = headers_proveedores
                
            for col in range(1, len(headers) + 1):
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col)
                
                for row_num in range(1, sheet.max_row + 1):
                    cell = sheet.cell(row=row_num, column=col)
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                sheet.column_dimensions[column_letter].width = max_length + 4

        # RESUMEN ESTADÍSTICO
        ws_resumen = wb.create_sheet("Resumen")
        ws_resumen.title = "Resumen"
        
        ws_resumen['A1'] = "RESUMEN ESTADÍSTICO - WINE SHOP"
        ws_resumen['A1'].font = title_font
        ws_resumen['A1'].alignment = align_center
        ws_resumen['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws_resumen['A2'].font = date_font
        ws_resumen['A2'].alignment = align_center

        # Estadísticas
        stats = [
            ("Total de Productos", Producto.objects.count()),
            ("Productos Activos", Producto.objects.filter(activo=True).count()),
            ("Productos en Stock", Producto.objects.filter(stock__gt=0).count()),
            ("Productos sin Stock", Producto.objects.filter(stock=0).count()),
            ("Productos en Oferta", Producto.objects.filter(precio_oferta__isnull=False).count()),
            ("Total de Proveedores", Proveedor.objects.count()),
            ("Movimientos de Inventario", Inventario.objects.count()),
            ("Valor Total del Inventario", f"${sum(p.precio * p.stock for p in Producto.objects.all()):.2f}"),
        ]

        # Escribir estadísticas
        for i, (label, value) in enumerate(stats, 4):
            ws_resumen.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws_resumen.cell(row=i, column=2, value=value)
            
            # Aplicar bordes y alineación
            for col in range(1, 3):
                cell = ws_resumen.cell(row=i, column=col)
                cell.border = border
                cell.alignment = align_center

        # Ajustar ancho de columnas del resumen
        for col in range(1, 3):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col)
            for row_num in range(1, ws_resumen.max_row + 1):
                cell = ws_resumen.cell(row=row_num, column=col)
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws_resumen.column_dimensions[column_letter].width = max_length + 4

        # Mover la hoja de resumen al principio
        wb.move_sheet("Resumen", 0)

        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Reporte_Inventario_WineShop_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'

        # Guardar el archivo
        wb.save(response)
        messages.success(request, "✅ Reporte de inventario completo generado exitosamente")
        return response

    except Exception as e:
        messages.error(request, f"❌ Error al generar el reporte: {str(e)}")
        return redirect('admin_panel')

@admin_required
def admin_panel(request):
    # Estadísticas generales
    usuarios = User.objects.count()
    productos = Producto.objects.count()
    pedidos = Pedido.objects.count()
    facturas = Factura.objects.count()
    inventario = Inventario.objects.count()
    clientes = Perfil.objects.filter(tipo_usuario='cliente').count()
    proveedores = Proveedor.objects.count()

    # Últimos pedidos
    ultimos_pedidos = Pedido.objects.select_related('cliente').order_by('-fecha_pedido')[:5]
    
    # Productos con bajo stock
    productos_bajo_stock = Producto.objects.filter(stock__lte=10).order_by('stock')[:5]
    
    # Últimos movimientos de inventario
    ultimos_movimientos = Inventario.objects.select_related('producto', 'responsable').order_by('-fecha')[:5]

    context = {
        'usuarios': usuarios,
        'productos': productos,
        'pedidos': pedidos,
        'facturas': facturas,
        'inventario': inventario,
        'clientes': clientes,
        'proveedores': proveedores,
        'ultimos_pedidos': ultimos_pedidos,
        'productos_bajo_stock': productos_bajo_stock,
        'ultimos_movimientos': ultimos_movimientos,
    }
    return render(request, 'admin_panel.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def inventario_view(request):
    movimientos = Inventario.objects.select_related('producto', 'responsable').order_by('-fecha')
    return render(request, 'inventario.html', {'movimientos': movimientos})






@admin_required
def productos_list(request):
    productos = Producto.objects.all().order_by('nombre')
    context = {'productos': productos}
    return render(request, 'crud/productos_list.html', context)


@admin_required
def producto_create(request):
    if request.method == 'POST':
        try:
            categoria_id = request.POST.get('categoria')
            proveedor_id = request.POST.get('proveedor')

            producto = Producto.objects.create(
                nombre=request.POST.get('nombre'),
                descripcion=request.POST.get('descripcion'),
                precio=request.POST.get('precio'),
                precio_oferta=request.POST.get('precio_oferta') or None,
                stock=request.POST.get('stock'),
                categoria_id=categoria_id,
                proveedor_id=proveedor_id or None,
                imagen_url=request.POST.get('imagen_url'),
                imagen_adicional_1=request.POST.get('imagen_adicional_1') or None,
                imagen_adicional_2=request.POST.get('imagen_adicional_2') or None,
                graduacion=request.POST.get('graduacion') or None,
                pais=request.POST.get('pais') or None,
                vina=request.POST.get('vina') or None,
                cosecha=request.POST.get('cosecha') or None,
                activo=request.POST.get('activo') == 'on'
            )
            messages.success(request, f'Producto "{producto.nombre}" creado exitosamente')
            return redirect('productos_list')
        except Exception as e:
            messages.error(request, f'Error al crear producto: {str(e)}')

    categorias = Categoria.objects.all()
    proveedores = Proveedor.objects.all()
    context = {'categorias': categorias, 'proveedores': proveedores}
    return render(request, 'crud/producto_form.html', context)


@admin_required
def producto_edit(request, id):
    producto = get_object_or_404(Producto, id=id)

    if request.method == 'POST':
        try:
            producto.nombre = request.POST.get('nombre')
            producto.descripcion = request.POST.get('descripcion')
            producto.precio = request.POST.get('precio')
            producto.precio_oferta = request.POST.get('precio_oferta') or None
            producto.stock = request.POST.get('stock')
            producto.categoria_id = request.POST.get('categoria')
            producto.proveedor_id = request.POST.get('proveedor') or None
            producto.imagen_url = request.POST.get('imagen_url')
            producto.imagen_adicional_1 = request.POST.get('imagen_adicional_1') or None
            producto.imagen_adicional_2 = request.POST.get('imagen_adicional_2') or None
            producto.graduacion = request.POST.get('graduacion') or None
            producto.pais = request.POST.get('pais') or None
            producto.vina = request.POST.get('vina') or None
            producto.cosecha = request.POST.get('cosecha') or None
            producto.activo = request.POST.get('activo') == 'on'
            producto.save()
            messages.success(request, f'Producto "{producto.nombre}" actualizado exitosamente')
            return redirect('productos_list')
        except Exception as e:
            messages.error(request, f'Error al actualizar producto: {str(e)}')

    categorias = Categoria.objects.all()
    proveedores = Proveedor.objects.all()
    context = {'producto': producto, 'categorias': categorias, 'proveedores': proveedores}
    return render(request, 'crud/producto_form.html', context)


@admin_required
def producto_delete(request, id):
    producto = get_object_or_404(Producto, id=id)

    if request.method == 'POST':
        nombre = producto.nombre
        producto.delete()
        messages.success(request, f'Producto "{nombre}" eliminado exitosamente')
        return redirect('productos_list')

    context = {'producto': producto}
    return render(request, 'crud/confirm_delete.html', context)


@admin_required
def proveedores_list(request):
    proveedores = Proveedor.objects.all().order_by('nombre')
    context = {'proveedores': proveedores}
    return render(request, 'crud/proveedores_list.html', context)


@admin_required
def proveedor_create(request):
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre')
            if not nombre:
                raise ValueError('El nombre es requerido')
                
            proveedor = Proveedor.objects.create(
                nombre=nombre,
                correo=request.POST.get('correo', ''),
                telefono=request.POST.get('telefono', ''),
                direccion=request.POST.get('direccion', '')
            )
            messages.success(request, f'Proveedor "{proveedor.nombre}" creado exitosamente')
            return redirect('proveedores_list')
        except Exception as e:
            messages.error(request, f'Error al crear proveedor: {str(e)}')

    return render(request, 'crud/proveedor_form.html', {'title': 'Crear Proveedor'})


@admin_required
def proveedor_edit(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)

    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre')
            if not nombre:
                raise ValueError('El nombre es requerido')
                
            proveedor.nombre = nombre
            proveedor.correo = request.POST.get('correo', '')
            proveedor.telefono = request.POST.get('telefono', '')
            proveedor.direccion = request.POST.get('direccion', '')
            proveedor.save()
            messages.success(request, f'Proveedor "{proveedor.nombre}" actualizado exitosamente')
            return redirect('proveedores_list')
        except Exception as e:
            messages.error(request, f'Error al actualizar proveedor: {str(e)}')

    context = {
        'proveedor': proveedor,
        'title': 'Editar Proveedor',
        'is_edit': True
    }
    return render(request, 'crud/proveedor_form.html', context)


@admin_required
def proveedor_delete(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)

    if request.method == 'POST':
        nombre = proveedor.nombre
        proveedor.delete()
        messages.success(request, f'Proveedor "{nombre}" eliminado exitosamente')
        return redirect('proveedores_list')

    context = {'proveedor': proveedor}
    return render(request, 'crud/confirm_delete.html', context)


@admin_required
def pedidos_list(request):
    pedidos = Pedido.objects.select_related('cliente').order_by('-fecha_pedido')
    context = {'pedidos': pedidos}
    return render(request, 'crud/pedidos_list.html', context)


@admin_required
def pedido_edit(request, id):
    pedido = get_object_or_404(Pedido, id=id)

    if request.method == 'POST':
        try:
            pedido.estado = request.POST.get('estado')
            pedido.direccion_envio = request.POST.get('direccion_envio')
            pedido.ciudad_envio = request.POST.get('ciudad_envio')
            pedido.telefono_envio = request.POST.get('telefono_envio')
            pedido.observaciones = request.POST.get('observaciones')
            pedido.save()
            messages.success(request, 'Pedido actualizado exitosamente')
            return redirect('pedidos_list')
        except Exception as e:
            messages.error(request, f'Error al actualizar pedido: {str(e)}')

    context = {'pedido': pedido}
    return render(request, 'crud/pedido_form.html', context)


@admin_required
def pedido_delete(request, id):
    pedido = get_object_or_404(Pedido, id=id)

    if request.method == 'POST':
        pedido_id = pedido.id
        pedido.delete()
        messages.success(request, f'Pedido #{pedido_id} eliminado exitosamente')
        return redirect('pedidos_list')

    context = {'pedido': pedido}
    return render(request, 'crud/confirm_delete.html', context)


@admin_required
def clientes_list(request):
    perfiles = Perfil.objects.filter(tipo_usuario='cliente').select_related('usuario').order_by('nombre_completo')
    context = {'clientes': perfiles}
    return render(request, 'crud/clientes_list.html', context)


@admin_required
def cliente_create(request):
    if request.method == 'POST':
        try:
            username = request.POST.get('username')
            if User.objects.filter(username=username).exists():
                messages.error(request, 'El nombre de usuario ya existe')
                return render(request, 'crud/cliente_form.html')

            user = User.objects.create_user(
                username=username,
                email=request.POST.get('email'),
                password=request.POST.get('password'),
                first_name=request.POST.get('nombre_completo').split()[0] if request.POST.get('nombre_completo') else ''
            )

            Perfil.objects.create(
                usuario=user,
                nombre_completo=request.POST.get('nombre_completo'),
                telefono=request.POST.get('telefono'),
                direccion=request.POST.get('direccion'),
                ciudad=request.POST.get('ciudad'),
                tipo_usuario='cliente'
            )
            messages.success(request, f'Cliente "{request.POST.get("nombre_completo")}" creado exitosamente')
            return redirect('clientes_list')
        except Exception as e:
            messages.error(request, f'Error al crear cliente: {str(e)}')

    return render(request, 'crud/cliente_form.html')


@admin_required
def cliente_edit(request, id):
    perfil = get_object_or_404(Perfil, id=id, tipo_usuario='cliente')

    if request.method == 'POST':
        try:
            perfil.nombre_completo = request.POST.get('nombre_completo')
            perfil.telefono = request.POST.get('telefono')
            perfil.direccion = request.POST.get('direccion')
            perfil.ciudad = request.POST.get('ciudad')
            perfil.save()

            perfil.usuario.email = request.POST.get('email')
            perfil.usuario.save()

            messages.success(request, 'Cliente actualizado exitosamente')
            return redirect('clientes_list')
        except Exception as e:
            messages.error(request, f'Error al actualizar cliente: {str(e)}')

    context = {'cliente': perfil}
    return render(request, 'crud/cliente_form.html', context)


@admin_required
def cliente_delete(request, id):
    perfil = get_object_or_404(Perfil, id=id, tipo_usuario='cliente')

    if request.method == 'POST':
        nombre = perfil.nombre_completo
        usuario = perfil.usuario
        perfil.delete()
        usuario.delete()
        messages.success(request, f'Cliente "{nombre}" eliminado exitosamente')
        return redirect('clientes_list')

    context = {'cliente': perfil}
    return render(request, 'crud/confirm_delete.html', context)


@admin_required
def usuarios_list(request):
    usuarios = User.objects.filter(is_superuser=True).order_by('username')
    context = {'usuarios': usuarios}
    return render(request, 'crud/usuarios_list.html', context)


@admin_required
def usuario_create(request):
    if request.method == 'POST':
        try:
            username = request.POST.get('username')
            if User.objects.filter(username=username).exists():
                messages.error(request, 'El nombre de usuario ya existe')
                return render(request, 'crud/usuario_form.html')

            user = User.objects.create_superuser(
                username=username,
                email=request.POST.get('email'),
                password=request.POST.get('password')
            )
            messages.success(request, f'Usuario administrador "{username}" creado exitosamente')
            return redirect('usuarios_list')
        except Exception as e:
            messages.error(request, f'Error al crear usuario: {str(e)}')

    return render(request, 'crud/usuario_form.html')


@admin_required
def usuario_edit(request, id):
    usuario = get_object_or_404(User, id=id, is_superuser=True)

    if request.method == 'POST':
        try:
            usuario.email = request.POST.get('email')
            usuario.first_name = request.POST.get('first_name')
            usuario.last_name = request.POST.get('last_name')
            usuario.save()
            messages.success(request, 'Usuario actualizado exitosamente')
            return redirect('usuarios_list')
        except Exception as e:
            messages.error(request, f'Error al actualizar usuario: {str(e)}')

    context = {'usuario': usuario}
    return render(request, 'crud/usuario_form.html', context)


@admin_required
def usuario_delete(request, id):
    usuario = get_object_or_404(User, id=id, is_superuser=True)

    if usuario.id == request.user.id:
        messages.error(request, 'No puedes eliminar tu propia cuenta')
        return redirect('usuarios_list')

    if request.method == 'POST':
        username = usuario.username
        usuario.delete()
        messages.success(request, f'Usuario "{username}" eliminado exitosamente')
        return redirect('usuarios_list')

    context = {'usuario': usuario}
    return render(request, 'crud/confirm_delete.html', context)


@admin_required
def roles_list(request):
    roles = Perfil.TIPOS_USUARIO
    perfiles_por_rol = {}
    for rol_value, rol_label in roles:
        perfiles_por_rol[rol_label] = Perfil.objects.filter(tipo_usuario=rol_value).count()

    context = {'roles': roles, 'perfiles_por_rol': perfiles_por_rol}
    return render(request, 'crud/roles_list.html', context)
